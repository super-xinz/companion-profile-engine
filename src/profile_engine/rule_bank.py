from __future__ import annotations

import hashlib
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook


DOMAIN_SHEETS = {"性格": "personality", "行为": "behavior", "做事工作": "work", "关系情感": "relationship"}
PROPORTION_SHEET = "比例"


@dataclass(frozen=True)
class RuleFragment:
    code: str
    domain: str
    source_field: str
    text: str
    normalized_weight: float
    source_sheet: str
    source_row: int
    source_column: int


def workbook_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


@lru_cache(maxsize=2)
def load_rule_index(workbook_path: str) -> dict[str, tuple[RuleFragment, ...]]:
    path = Path(workbook_path)
    if not path.exists():
        return {}
    wb = load_workbook(path, read_only=True, data_only=True)
    index: dict[str, list[RuleFragment]] = {}
    seen: set[tuple] = set()
    text_pool: dict[str, str] = {}
    try:
        for sheet_name, domain in DOMAIN_SHEETS.items():
            ws = wb[sheet_name]
            headers = None
            for row_no, row in enumerate(ws.iter_rows(values_only=True), 1):
                if headers is None:
                    if len(row) > 1 and row[1] == "打头数字":
                        headers = list(row)
                    continue
                raw_code = row[0]
                if raw_code is None:
                    continue
                normalized_code = str(raw_code).split(".")[0].zfill(4)
                for text_column in range(2, len(row), 2):
                    text = row[text_column - 1]
                    weight = row[text_column] if text_column < len(row) else None
                    field = headers[text_column - 1]
                    if text is None or not str(text).strip() or str(text).strip() == "42":
                        continue
                    try:
                        normalized_weight = float(weight or 0)
                    except (TypeError, ValueError):
                        continue
                    if normalized_weight <= 0:
                        continue
                    clean_text_value = " ".join(str(text).split())
                    clean_text = text_pool.setdefault(clean_text_value, clean_text_value)
                    key = (normalized_code, domain, field, clean_text)
                    if key in seen:
                        continue
                    seen.add(key)
                    index.setdefault(normalized_code, []).append(RuleFragment(code=normalized_code, domain=domain, source_field=str(field), text=clean_text,
                        normalized_weight=normalized_weight, source_sheet=sheet_name, source_row=row_no, source_column=text_column))
    finally:
        wb.close()
    return {code: tuple(fragments) for code, fragments in index.items()}


@lru_cache(maxsize=2)
def load_domain_proportions(workbook_path: str) -> dict[str, tuple[float, ...]]:
    path = Path(workbook_path)
    if not path.exists():
        return {}
    wb = load_workbook(path, read_only=True, data_only=True)
    proportions: dict[str, tuple[float, ...]] = {}
    label_map = {
        "性格": "personality",
        "行为": "behavior",
        "做事/事业": "work",
        "关系/感情": "relationship",
    }
    try:
        ws = wb[PROPORTION_SHEET]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            return {}
        for row_no, row in enumerate(rows):
            label = str(row[0] or "").strip()
            domain = label_map.get(label)
            if not domain:
                continue
            ratios: list[float] = []
            ratio_row = rows[row_no + 1] if row_no + 1 < len(rows) else ()
            for value in ratio_row:
                if value in (None, ""):
                    continue
                try:
                    ratios.append(float(value))
                except (TypeError, ValueError):
                    ratios.append(0.0)
            proportions[domain] = tuple(ratios)
    finally:
        wb.close()
    return proportions


def fragments_for_code(workbook_path: str, code: str) -> tuple[RuleFragment, ...]:
    return load_rule_index(workbook_path).get(code, ())


def extract_signals(fragments: tuple[RuleFragment, ...], cold_rules: dict) -> list[dict]:
    dictionary = cold_rules["semantic_signal_extraction"]["generalized_signal_dictionary"]
    signals: list[dict] = []
    for fragment in fragments:
        for signal_id, rule in dictionary.items():
            matched = [cue for cue in rule.get("cues", []) if cue and cue in fragment.text]
            if not matched:
                continue
            for target, direction in rule.get("effects", {}).items():
                signals.append({
                    "signal_id": signal_id, "target": target, "direction": int(direction),
                    "strength": min(1.0, max(0.0, fragment.normalized_weight * 4)),
                    "normalized_weight": fragment.normalized_weight, "matched_cues": matched,
                    "source_fragment": {"domain": fragment.domain, "field": fragment.source_field,
                        "sheet": fragment.source_sheet, "row": fragment.source_row, "column": fragment.source_column,
                        "text_sha256": hashlib.sha256(fragment.text.encode()).hexdigest()},
                })
    return signals
