from __future__ import annotations

import re
from datetime import datetime, timezone
from pathlib import Path
from copy import deepcopy
from typing import Any

from openpyxl import load_workbook

from .template_people import TEMPLATE_BY_BIRTH_DATE, template_person_for_birth_date


SOURCE_FILES = {birth_date: person.source_file for birth_date, person in TEMPLATE_BY_BIRTH_DATE.items()}
SOURCE_META = {
    birth_date: {**person.birth_analysis, "mbti": person.mbti}
    for birth_date, person in TEMPLATE_BY_BIRTH_DATE.items()
}


SOURCE_PROFILE_SCHEMA_VERSION = "source-profile-structured-v2"

BEHAVIOR_SCENARIOS = {
    "接到任务": ("task_style", "task_received"),
    "推进过程": ("task_style", "task_progress"),
    "遇到阻碍": ("task_style", "obstacle"),
    "做决策": ("task_style", "decision"),
    "被催促": ("task_style", "being_urged"),
    "出错后": ("task_style", "after_error"),
    "面对变化": ("task_style", "facing_change"),
    "初识阶段": ("relationship_style", "first_meeting"),
    "熟识后": ("relationship_style", "familiar_relationship"),
    "帮助别人": ("relationship_style", "helping_others"),
    "被需要时": ("relationship_style", "being_needed"),
    "被误解时": ("relationship_style", "being_misunderstood"),
    "冲突中": ("relationship_style", "conflict"),
    "对异性": ("relationship_style", "romantic_interaction"),
    "自信": ("inner_state_style", "confidence_state"),
    "乐观/悲观": ("inner_state_style", "optimism_state"),
    "压力状态": ("inner_state_style", "stress_response"),
    "能量来源": ("inner_state_style", "energy_source"),
}

LANGUAGE_SECTION_NAMES = {
    "说话方式": "speaking_style",
    "幽默感": "humor",
    "情绪表达": "emotion_expression",
    "典型话术": "typical_utterances",
    "极少说的话": "rare_utterances",
}

TYPICAL_UTTERANCE_KEYS = {
    "被征求意见": "asked_for_opinion",
    "被夸": "praised",
    "被批评": "criticized",
    "表达不满": "expressing_dissatisfaction",
    "说不": "saying_no",
    "安慰人": "comforting",
    "分享好消息": "sharing_good_news",
    "紧张时": "nervous",
    "放松时": "relaxed",
}

LANGUAGE_HEADER_LABELS = {"特征", "类型", "情境", "表现", "举例", "ta大概会说的话"}


def _source_path(birth_date: str) -> Path | None:
    person = template_person_for_birth_date(birth_date)
    if not person:
        return None
    filename = person.source_file
    project_root = Path(__file__).resolve().parents[2]
    candidates = [
        project_root / "source_profiles" / filename,
        Path.cwd() / "source_profiles" / filename,
        Path.cwd() / filename,
        Path.cwd().parent / filename,
    ]
    return next((path for path in candidates if path.exists()), None)


def _rows(sheet) -> list[list[Any]]:
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    while rows and not any(value is not None for value in rows[-1]):
        rows.pop()
    width = max((len(row) for row in rows), default=0)
    while width and all((row[width - 1] if len(row) >= width else None) is None for row in rows):
        width -= 1
    return [row[:width] for row in rows]


def _cell_text(value: Any) -> str:
    return str(value or "").strip()


def _normalized_source_label(value: Any) -> str:
    label = _cell_text(value).replace('"', "").replace("“", "").replace("”", "")
    if label.startswith("对异性"):
        return "对异性"
    if label.endswith("需要时"):
        return "被需要时"
    if label.startswith("说") and "不" in label:
        return "说不"
    return label


def parse_source_behavior(document: dict) -> dict[str, dict[str, dict[str, Any]]]:
    """Turn the workbook's scenario table into the canonical 18-scenario structure."""
    result: dict[str, dict[str, dict[str, Any]]] = {
        "task_style": {}, "relationship_style": {}, "inner_state_style": {},
    }
    for row in document.get("sheets", {}).get("行为风格", []):
        if len(row) < 2:
            continue
        source_label = _normalized_source_label(row[0])
        target = BEHAVIOR_SCENARIOS.get(source_label)
        if not target or not row[1]:
            continue
        group_key, scenario_key = target
        result[group_key][scenario_key] = {
            "source_label": source_label,
            "feature": _cell_text(row[1]),
            "parameter_text": _cell_text(row[2]) if len(row) > 2 else "",
            "explanation": _cell_text(row[3]) if len(row) > 3 else "",
            "confidence": 0.45,
            "origin": "user_supplied_complete_profile",
        }
    imported = {key for scenarios in result.values() for key in scenarios}
    expected = {scenario_key for _, scenario_key in BEHAVIOR_SCENARIOS.values()}
    missing = sorted(expected - imported)
    if missing:
        raise ValueError(f"{document.get('source_file', '原始画像工作簿')} 缺少行为场景: {missing}")
    return result


def parse_source_language(document: dict) -> dict[str, Any]:
    """Preserve the workbook's language examples as structured baseline content."""
    result: dict[str, Any] = {
        "speaking_style": [], "humor": [], "emotion_expression": [],
        "typical_utterances": {}, "rare_utterances": [],
    }
    current_section: str | None = None
    for row in document.get("sheets", {}).get("语言风格", []):
        if not row:
            continue
        label = _cell_text(row[0])
        heading = next((key for source_name, key in LANGUAGE_SECTION_NAMES.items()
                        if source_name in label and "、" in label), None)
        if heading:
            current_section = heading
            continue
        normalized_label = _normalized_source_label(label)
        if (not current_section or not normalized_label
                or normalized_label in LANGUAGE_HEADER_LABELS):
            continue
        behavior = _cell_text(row[1]) if len(row) > 1 else ""
        example = _cell_text(row[2]) if len(row) > 2 else ""
        if current_section in {"speaking_style", "humor", "emotion_expression"}:
            result[current_section].append({
                "label": normalized_label,
                "behavior": behavior,
                "example": example or None,
                "confidence": 0.45,
                "evidence_refs": [],
                "origin": "user_supplied_complete_profile",
            })
        elif current_section == "typical_utterances":
            key = TYPICAL_UTTERANCE_KEYS.get(normalized_label)
            if key and behavior:
                result[current_section][key] = {
                    "label": normalized_label,
                    "utterance_pattern": behavior,
                    "example": behavior,
                    "parameter_refs": [],
                    "evidence_refs": [],
                    "confidence": 0.45,
                    "origin": "user_supplied_complete_profile",
                }
        elif current_section == "rare_utterances" and behavior:
            result[current_section].append({
                "utterance_or_pattern": normalized_label,
                "reason": behavior,
                "evidence_refs": [],
                "confidence": 0.45,
                "origin": "user_supplied_complete_profile",
            })
    if not result["speaking_style"] or not result["typical_utterances"]:
        raise ValueError(f"{document.get('source_file', '原始画像工作簿')} 缺少语言风格主体内容")
    return result


def hydrate_source_sections(profile: dict, document: dict | None = None) -> bool:
    """Attach structured behavior/language fixtures without replacing dialogue evidence."""
    document = document or profile.get("source_profile_document")
    if not isinstance(document, dict):
        return False

    source_behavior = parse_source_behavior(document)
    profile["source_behavior"] = deepcopy(source_behavior)
    behavior_style = profile.setdefault("behavior_style", {})
    for group_key, scenarios in source_behavior.items():
        current_group = behavior_style.setdefault(group_key, {})
        for scenario_key, source_item in scenarios.items():
            current = current_group.setdefault(scenario_key, {})
            direct_refs = list(current.get("direct_evidence_refs", []))
            observations = list(current.get("observations", []))
            parameter_refs = list(current.get("parameter_refs", []))
            confidence = max(float(current.get("confidence", 0)), source_item["confidence"])
            current.update({
                **source_item,
                "parameter_refs": parameter_refs,
                "direct_evidence_refs": direct_refs,
                "observations": observations,
                "generation_rule_id": f"SOURCE-WORKBOOK-{scenario_key}",
                "confidence": confidence,
            })

    source_language = parse_source_language(document)
    missing_typical = set(TYPICAL_UTTERANCE_KEYS.values()) - set(source_language["typical_utterances"])
    if "nervous" in missing_typical:
        trait_lookup = {
            key: item for category in profile.get("core_traits", {}).values()
            for key, item in category.items()
        }
        impulsivity = float(trait_lookup.get("impulsivity", {}).get("value", .5))
        if impulsivity >= .65:
            pattern = "紧张时语速可能加快，但仍会努力把结论和理由表达清楚。"
        elif impulsivity <= .30:
            pattern = "紧张时会减少表达、放慢语速，并更谨慎地选择措辞。"
        else:
            pattern = "紧张时会收紧表达范围，优先确认事实和下一步行动。"
        source_language["typical_utterances"]["nervous"] = {
            "label": "紧张时",
            "utterance_pattern": pattern,
            "example": pattern,
            "parameter_refs": ["impulsivity"],
            "evidence_refs": [],
            "confidence": 0.45,
            "origin": "derived_from_complete_profile",
        }
    profile["source_language"] = deepcopy(source_language)
    previous_language = profile.get("language_style", {})
    observed = [
        item for item in previous_language.get("speaking_style", [])
        if item.get("origin") == "observed"
    ]
    profile["language_style"] = deepcopy(source_language)
    if observed:
        profile["language_style"]["speaking_style"] = [
            *observed[-6:], *profile["language_style"]["speaking_style"],
        ]
    if previous_language.get("observation_state"):
        profile["language_style"]["observation_state"] = deepcopy(
            previous_language["observation_state"]
        )
    profile.setdefault("meta", {})["source_profile_schema_version"] = SOURCE_PROFILE_SCHEMA_VERSION
    return True


def load_source_document(birth_date: str) -> dict | None:
    path = _source_path(birth_date)
    if not path:
        return None
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return {
            "birth_date": birth_date,
            "source_file": path.name,
            "source_type": "user_supplied_complete_profile",
            "imported_at": datetime.now(timezone.utc).isoformat(),
            "sheets": {name: _rows(workbook[name]) for name in workbook.sheetnames},
        }
    finally:
        workbook.close()


def apply_source_profile(profile: dict, birth_date: str) -> bool:
    """Overlay the exact user-supplied profile and retain every original workbook cell."""
    document = load_source_document(birth_date)
    if not document:
        return False

    overview = document["sheets"].get("性格总览", [])
    meta = SOURCE_META[birth_date]
    profile["birth_analysis"].update({
        "bazi_text": meta["bazi_text"],
        "day_master": meta["day_master"],
        "pattern_name": meta["pattern_name"],
        "strength_label": meta["strength_label"],
        "relation_markers": meta["relation_markers"],
    })

    mbti_labels = {"E↔I": "ei", "S↔N": "sn", "T↔F": "tf", "J↔P": "jp"}
    imported_mbti = set()
    for row in overview:
        label = str(row[0] or "").replace(" ", "")
        key = next((value for prefix, value in mbti_labels.items() if label.startswith(prefix)), None)
        if not key or len(row) < 5:
            continue
        profile["mbti_dimensions"][key].update({
            "value": float(row[1]),
            "confidence": 0.45,
            "source_tendency": row[2],
            "source_interpretation": row[4],
        })
        imported_mbti.add(key)
    missing_mbti = set(mbti_labels.values()) - imported_mbti
    if missing_mbti:
        raise ValueError(f"{document['source_file']} 缺少MBTI维度: {sorted(missing_mbti)}")
    profile["mbti_dimensions"]["type_label"] = meta["mbti"]

    trait_lookup = {
        key: value
        for category in profile["core_traits"].values()
        for key, value in category.items()
    }
    imported_traits = set()
    for row in overview:
        if len(row) < 6:
            continue
        match = re.match(r"^([a-z_]+)", str(row[1] or ""))
        if not match or match.group(1) not in trait_lookup:
            continue
        key = match.group(1)
        tendency = str(row[4] or "").replace("⭐", "").strip()
        interpretation = str(row[5] or "").strip()
        trait_lookup[key].update({
            "value": float(row[2]),
            "confidence": 0.45,
            "tendency_label": tendency or trait_lookup[key]["tendency_label"],
            "interpretation": interpretation or tendency,
            "origin": "user_supplied_complete_profile",
        })
        imported_traits.add(key)
    missing_traits = set(trait_lookup) - imported_traits
    if missing_traits:
        raise ValueError(f"{document['source_file']} 缺少核心维度: {sorted(missing_traits)}")

    portrait_labels = {
        "本质": "essence",
        "优势": "strengths",
        "弱点": "weaknesses",
        "核心矛盾": "core_tension",
        "适合角色": "suitable_roles",
        "适合的角色": "suitable_roles",
    }
    source_portrait = {}
    for row in document["sheets"].get("人物画像", []):
        if len(row) < 2:
            continue
        key = portrait_labels.get(str(row[0] or "").strip())
        if key and row[1]:
            source_portrait[key] = {
                "content": row[1],
                "confidence": 0.45,
                "origin": "user_supplied_complete_profile",
            }
    profile["source_portrait"] = source_portrait
    missing_portrait = set(portrait_labels.values()) - set(source_portrait)
    if missing_portrait:
        raise ValueError(f"{document['source_file']} 缺少人物画像字段: {sorted(missing_portrait)}")
    profile["portrait"] = {
        key: {
            "content": item["content"],
            "confidence": 0.45,
            "origin": "user_supplied_complete_profile",
            "parameter_refs": [],
        }
        for key, item in source_portrait.items()
    }
    profile["source_profile_document"] = document
    hydrate_source_sections(profile, document)
    profile["identity"]["template_person_id"] = template_person_for_birth_date(birth_date).user_id
    profile["meta"]["warnings"] = [
        warning for warning in profile["meta"].get("warnings", [])
        if "黄金样例" not in warning and "专家尚未提供任意生日到四位数字学编码" not in warning
    ]
    profile["meta"]["warnings"].append("已完整导入用户提供的原始画像工作簿；后续对话证据将持续校准动态画像。")
    return True
